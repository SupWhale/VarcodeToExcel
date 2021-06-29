from PIL import Image
from PIL import ImageTk
import keyboard
import tkinter as tk
import threading
import datetime
import os
import time
import cv2
import pyzbar.pyzbar as pyzbar
from openpyxl import Workbook
from datetime import datetime


def decode_serial(im): #이미지내에서 바코드를 찾아내고 해당 타입과 데이터를 출력하는 함수
    
    decodedObjects = pyzbar.decode(im)

    result = ""
    for obj in decodedObjects:
        print('Type : ', obj.type)
        print('Data : ', obj.data, '\n')
        result = obj.data
        print(result[0])
    return result

def decode(im): #이미지내에서 바코드를 찾아내고 해당 타입과 데이터를 출력하는 함수
    
    decodedObjects = pyzbar.decode(im)

    for obj in decodedObjects:
        print('Type : ', obj.type)
        print('Data : ', obj.data, '\n')
    return decodedObjects


def macadd(mac):
    
    result = ''
    
    if len(mac) == 12:
        
        list = [0,0,0,0,0,0,0,0,0,0,0]
        
        for i in range(0,11):
            if i%2 == 0:
                list[i] = mac[i:i+2]
            elif i%2 != 0:
                list[i] = ':'
        
        for j in list:
            result += j
    else:
        result = mac
    return result

def CheckType_Serial(code):
    try:   
        result = ""
        if code[0] == 84:
            result = "TMS"
        elif code[0] == 77:
            result = "PMC"
        elif code[0] == 71:
            result = "GateWay"
        elif code[0] == 82:
            result = "IRC"
        else:
            result = ""
    except IndexError as e:
        print(e)
    
    return result
        
def Setmode_Doc(ch):
    global choice
    State.config(text = '의사협회모드')
    choice = ch + 1
    
def Setmode_Stock(ch):
    global choice
    State.config(text = '재고관리모드')
    choice = ch + 2
    
def Setmode_Install(ch):
    global choice
    State.config(text = '설치정보모드')
    choice = ch + 3

def Doctor(write_ws, decodedObjects,width,high): #의사협회 관리 모드
    
    for obj in decodedObjects:
        code = obj.data
        result.config(text = obj.data)
        write_ws.cell(high,width,macadd(code.decode()))     
       
def Stock_Manage(write_ws, decodedObjects,width,high): #재고 관리 모드

        result.config(text = decodedObjects)
        write_ws.cell(high,width,decodedObjects)
        write_ws.cell(high,width+1,'회사 7층')

        stri = datetime.today().strftime("%Y-%m-%d")
        
        write_ws.cell(high,width+2,stri)

def install_Int(write_ws, decodedObjects,width,high): #GEM 설치정보 양식에 맞춘 모드
    
    for obj in decodedObjects:
        code = obj.data
        result.config(text = obj.data)
        str_high = str(high)
        location = 'A'+str_high
        write_ws[location] = obj.data
        merge_location = 'A'+str_high+':C'+str_high
        write_ws.merge_cells(merge_location)
    
def camThread():
    
    cap = cv2.VideoCapture(0) #외부 웹캠을 불러옴
    write_wb = Workbook() #워크시트를 생성할 객체 생성
    panel = None #GUI를 생성할 판넬 생성
    write_ws = write_wb.create_sheet('기본시트') #워크 시트 생성
    gateway_ws = write_wb.create_sheet('Gateway')
    pmc_ws = write_wb.create_sheet('PMC')
    tms_ws = write_wb.create_sheet('TMS')
    irc_ws = write_wb.create_sheet('IRC')
    
    print('width :%d, height : %d' % (cap.get(3), cap.get(4))) #현재 카메라의 해상도 출력

    width = 2
    high = 3
    pmc_h=1
    irc_h=1
    tms_h=1
    gateway_h=1
    check_code = False
    while(True):
        ret, frame = cap.read()    # Read 결과와 frame
        if(ret) :
            
            image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            image = Image.fromarray(image)
            image = ImageTk.PhotoImage(image)

            if panel is None:
                panel = tk.Label(image=image)
                panel.image = image
                panel.pack(side="left")
                
            else:
                panel.configure(image=image)
                panel.image = image
            
            if keyboard.is_pressed('q'):
                time.sleep(0.1)
                cv2.imwrite('bacode.jpg', frame) #이미지파일 출력
                im = cv2.imread('bacode.jpg')
                    
                decodedObjects = decode(im)#바코드 해석 
                print(decodedObjects)
              
                if decodedObjects == '[]':
                    print("감지된 바코드가 없습니다.")
                    continue
                
                else:
                
                    if choice == 1:
                        
                        Doctor(write_ws, decodedObjects,width,high)
                       
                        if width < 6:
                            width = width+1
                        elif width == 6:
                            width = 3
                            high = high+1
                            
                    elif choice == 2:
                        
                        decodedObjects_s = decode_serial(im)
                        print(decodedObjects_s)

                        types = CheckType_Serial(decodedObjects_s)
                        
                        if types == "TMS":
                            
                            #tms_ws = write_wb.active
                            Stock_Manage(tms_ws, decodedObjects_s,width,tms_h)
                            tms_h = tms_h+1
                            
                        elif types == "PMC":
                            
                            Stock_Manage(pmc_ws, decodedObjects_s,width,pmc_h)
                            pmc_h = pmc_h + 1
                            
                        elif types == "IRC":

                            Stock_Manage(irc_ws, decodedObjects_s,width,irc_h)
                            irc_h = irc_h + 1
                            
                        elif types == "GateWay":
                            
                            Stock_Manage(gateway_ws, decodedObjects_s,width,gateway_h)
                            gateway_h = gateway_h+1
                            
                        high = high+1
                        
                    elif choice == 3:
                        
                        install_Int(write_ws, decodedObjects,width,high)
                        high = high+1
                    
                    str = datetime.today().strftime("%Y년%m월%d일%H시")
                        
                    write_wb.save(str + '.xlsx') #엑셀 파일 저장
                
    cap.release() #프로그램 최종종료
    cv2.destroyAllWindows()
       
if __name__ == '__main__': #Main
    thread_img = threading.Thread(target=camThread, args=())
    thread_img.daemon = True
    thread_img.start()

    ch = 0
    
    root = tk.Tk()
    root.geometry("1200x600+100+100")

    State = tk.Label(root, text = '모드를 입력하십시오', font = 'TkFixedFont')
    State.pack()

    button1 = tk.Button(root, overrelief="solid", text = "의사협회 모드" ,width=15, command = lambda: Setmode_Doc(ch))
    button1.place(x=1000, y=500)

    button2 = tk.Button(root, overrelief="solid", text = "재고관리 모드" ,width=15, command = lambda: Setmode_Stock(ch))
    button2.place(x=1000, y=400)

    button3 = tk.Button(root, overrelief="solid", text = "설치정보 모드" ,width=15, command = lambda: Setmode_Install(ch))
    button3.place(x=1000, y=300)

    result = tk.Label(root, text = 'Result', font = 'TkFixedFont')
    result.pack()
    
    root.mainloop()

    
    


    

