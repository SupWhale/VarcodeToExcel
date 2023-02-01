from PIL import Image
from PIL import ImageTk
import keyboard
import tkinter as tk
import threading
import datetime
import os
import time
import cv2
import pyzbar.pyzbar as pyzbar
from openpyxl import *
from datetime import datetime
from toExcel import *
from decode import *
from GEM import *
import sys


def Setmode_Doc(ch):
    global choice  #모드 변경을 식별해주는 상수
    State.config(text = '의사협회모드') #제목 창 이름 변경
    choice = ch + 1
    
def Setmode_Stock(ch):
    global choice
    State.config(text = '재고관리모드')
    choice = ch + 2
    
def Setmode_Install(ch):
    global choice
    State.config(text = '설치정보모드')
    choice = ch + 3

def Save_Filename(types):
    StorageFile = "" #저장할 파일 명
    
    if types == 2:
        StorageFile = "재고 관리 파일"

    else:
        StorageFile = datetime.today().strftime("%Y년%m월%d일%H시") #오늘 날짜로 파일명 저장
        
    return StorageFile

def camThread():
    ch = 0 #모드 변경 식별 상수
    button1 = tk.Button(root, overrelief="solid", text = "의사협회 모드", width=15, command = lambda: Setmode_Doc(ch))
    button1.place(x=1000, y=500)

    button2 = tk.Button(root, overrelief="solid", text = "재고관리 모드", width=15, command = lambda: Setmode_Stock(ch))
    button2.place(x=1000, y=400)

    button3 = tk.Button(root, overrelief="solid", text = "설치정보 모드", width=15, command = lambda: Setmode_Install(ch))
    button3.place(x=1000, y=300)

    try:
            cap = cv2.VideoCapture(1) #외부 웹캠을 불러옴
            Write_WB = Workbook() #워크시트를 생성할 객체 생성
            Write_WS = Write_WB.create_sheet('기본시트') #워크 시트 생성
            panel = None #GUI를 생성할 판넬 생성
            lable_WH = [2,2,3,1,1,1,1] #width2(0), width(1)(작성이 시작될 X 좌표), high(2)(작성이 시작될 Y 좌표), 
                                       #pmc_h(3) 작성할 Y좌표, irc_h(4) 작성할 Y좌표, tms_h(5) 작성할 Y좌표, gateway_h(6) 작성할 Y좌표
                    
            print('width : %d, height : %d' % (cap.get(3), cap.get(4))) #현재 카메라의 해상도 출력

            check_code = False
            
            while(True):
                end_point = False
                ret, frame = cap.read()    # Read 결과와 frame
                if(ret) :
                    image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    image = Image.fromarray(image)
                    image = ImageTk.PhotoImage(image)

                    if panel is None:
                        panel = tk.Label(image=image)
                        panel.image = image
                        panel.pack(side="left")
    
                    else:
                        panel.configure(image=image)
                        panel.image = image
                    
                    if choice == 2: #재고 관리 모드일 경우
                            Write_WB = load_workbook("재고 관리 파일.xlsx") #워크북 작성
                            PMC_WS = Write_WB['PMC']        #장비별 워크시트 작성
                            Gateway_WS = Write_WB['Gateway']
                            TMS_WS = Write_WB['TMS']
                            IRC_WS = Write_WB['IRC']

                            lable_WH[1] = 3
                            lable_WH[2] = 3
                            lable_WH[3] = PMC_WS.cell(3,20).value #엑셀 파일에 기재되어있는 재고 개수를 기반으로 Y 좌표 수정
                            lable_WH[4] = IRC_WS.cell(3,20).value
                            lable_WH[5] = TMS_WS.cell(3,20).value
                            lable_WH[6] = Gateway_WS.cell(3,20).value

                    else:   #데이터베이스 저장 모드
                        GEM_Database = load_workbook("GEM 데이터베이스.xlsx") 
                        GEM_WS = GEM_Database['Sheet']
                        DB_High = GEM_WS.cell(2,20).value

                    if keyboard.is_pressed('q'):    #이미지 촬영모드
                        time.sleep(0.1)
                        cv2.imwrite('bacode.jpg', frame) #이미지파일 출력
                        im = cv2.imread('bacode.jpg')
                            
                        DecodeObject = Decode_Image(im) #바코드 해석 
                        result.config(text = Decode_Serial(im))
                        
                        if DecodeObject == []:
                            result.config(text = "감지된 바코드가 없습니다.")
                            continue

                        else: #설정된 모드에 따라 적절한 문서 양식으로 문서를 작성합니다.
                        
                            if choice == 1:    #의사협회 정보 관리 모드
                                DB_Sell(GEM_WS, DecodeObject, lable_WH[0], DB_High) #문서 정보 불러오기

                                if lable_WH[0] < 3:
                                    lable_WH[0] = lable_WH[0]+1

                                elif lable_WH[0] == 3:
                                    lable_WH[0] = 2
                                    DB_High = DB_High+1

                                Doctor_Processing(Write_WS, DecodeObject, lable_WH[1], lable_WH[2])
                                
                                if lable_WH[1] < 7:
                                    lable_WH[1] = lable_WH[1]+1

                                elif lable_WH[1] == 7:
                                    lable_WH[1] = 3
                                    lable_WH[2] = lable_WH[2]+1
                                                                    
                            elif choice == 2: #재고 관리 모드
                                DecodedObjects_s = Decode_Serial(im)
                                print(DecodedObjects_s)

                                types = CheckType_Serial(DecodedObjects_s)
                                
                                if types == "TMS":   
                                    Stock_Manage(TMS_WS, DecodeObject, lable_WH[1], lable_WH[5])
                                    lable_WH[5] = lable_WH[5]+1
                                    
                                elif types == "PMC":
                                    Stock_Manage(PMC_WS, DecodeObject, lable_WH[1], lable_WH[3])
                                    lable_WH[3] = lable_WH[3]+ 1
                                    
                                elif types == "IRC":
                                    Stock_Manage(IRC_WS, DecodeObject, lable_WH[1], lable_WH[4])
                                    lable_WH[4] = lable_WH[4]+ 1
                                    
                                elif types == "GateWay":  
                                    Stock_Manage(Gateway_WS, DecodeObject, lable_WH[1], lable_WH[6])
                                    lable_WH[6] = lable_WH[6]+1
                                    
                                lable_WH[2] = lable_WH[2]+1
                                
                            elif choice == 3: #GEM 설치정보 양식 작성 모드
                                DB_Sell(GEM_WS, DecodeObject, lable_WH[0], DB_High) #문서 정보 불러오기

                                if lable_WH[0] < 3:
                                    lable_WH[0] = lable_WH[0]+1

                                elif lable_WH[0] == 3:
                                    lable_WH[0] = 2
                                    DB_High = DB_High+1

                                Install_Int(Write_WS, DecodeObject, lable_WH[2])
                                lable_WH[2] = lable_WH[2]+1
                                                        
                            FileName = Save_Filename(choice)                           
                            Write_WB.save(FileName + '.xlsx') #엑셀 파일 저장
                        
            cap.release() #프로그램 최종종료
            cv2.destroyAllWindows() 
    except NameError as e:
        print(e)
        result.config(text = "잠시후 프로그램이 재시작됩니다.")
        ErrorMessage.config(text = e)
        time.sleep(15)
        os.excel(sys.executable, sys.executable, *sys.argv)

    except PermissionError as e:
        print(e)
        result.config(text = "현재 활성화 된 엑셀 파일을 닫아주세요! 잠시후 프로그램이 재시작됩니다.")
        ErrorMessage.config(text = e)
        time.sleep(15)
        os.excel(sys.executable, sys.executable, *sys.argv)
        
if __name__ == '__main__': #Main
    try:         
            root = tk.Tk()
            root.title("GEM 자동화 엑셀 작성 툴")
            root.geometry("1200x600+100+100")

            State = tk.Label(root, text = '모드를 입력하십시오', font = 'TkFixedFont')
            State.pack()

            result = tk.Label(root, text = 'Result', font = 'TkFixedFont')
            result.pack()

            ErrorMessage = tk.Label(root, text = 'Result', font = 'TkFixedFont')
            ErrorMessage.pack()
            
            thread_img = threading.Thread(target=camThread, args=())
            thread_img.daemon = True
            thread_img.start()
            
            root.mainloop()

    except NameError as e:
        print(e)
