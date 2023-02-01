from PIL import Image
from PIL import ImageTk
import keyboard
import tkinter as tk
import threading
import datetime
import os
import time
import cv2
import pyzbar.pyzbar as pyzbar
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from GEM import *

def Mac_Address(mac): #맥어드레스 추출 함수
    result = ''
    if len(mac) == 12: #제품번호 총 12자리
        BacodeList = [0,0,0,0,0,0,0,0,0,0,0]
        
        for i in range(0,11):
            if i%2 == 0:
                BacodeList[i] = mac[i:i+2] #맥어드레스 추출
            elif i%2 != 0:
                BacodeList[i] = ':'
        
        for j in list:
            result += j
    else:
        result = mac
    return result

def CheckType_Serial(ProductCode): #장비 분류 함수
    try:   
        result = ""

        if ProductCode[0] == 84:
            result = "TMS"

        elif ProductCode[0] == 77:
            result = "PMC"

        elif ProductCode[0] == 71:
            result = "GateWay"

        elif ProductCode[0] == 82:
            result = "IRC"

        else:
            result = ""
            
    except IndexError as e:
        print(e)
    
    return result
        
def Doctor_Processing(Write_WS, DecodedObjects,width,high): #의사협회 관리 모드
    for obj in DecodedObjects:
        code = obj.data
        Write_WS.cell(high,width,Mac_Address(code.decode()))     
       
def Stock_Manage(Write_WS, DecodedObjects,width,high): #재고 관리 모드
    for obj in DecodedObjects:
        ProductCode = obj.data #해석된 시리얼 번호를 가져옴
        Write_WS.cell(high,width,ProductCode)
        Write_WS.cell(high,width+1,'회사 7층')

        Datetime = datetime.today().strftime("%Y-%m-%d")
        color =  PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')

        Write_WS.cell(high,width+2,Datetime)
        Write_WS.cell(2,20, '아래 셀을 절대 지우지 마세요').fill = color
        Write_WS.cell(2,20,high).fill = color

def Install_Int(Write_WS, DecodedObjects, high): #GEM 설치정보 양식에 맞춘 모드
    for obj in DecodedObjects:
        str_high = str(high)
        location = 'A'+str_high
        Write_WS[location] = obj.data
        MergeLocatoin = 'A'+str_high+':C'+str_high
        Write_WS.merge_cells(MergeLocatoin)
    
def DB_Sell(Write_WS, DecodedObjects, width, high): #해석한 바코드 데이터를 바탕으로 문서를 새로 작성합니다
    for obj in DecodedObjects:
        ProductCode = obj.data
        Write_WS.cell(high,width,Mac_Address(ProductCode.decode()))

        Datetime = datetime.today().strftime("%Y-%m-%d")
        color =  PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')

        Write_WS.cell(high,width+2,Datetime)
        Write_WS.cell(1,20, '아래 셀을 절대 지우지 마세요').fill = color
        Write_WS.cell(2,20,high).fill = color

    
